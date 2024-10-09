import gspread  # імпортуєм бібліотеки
import csv
from datetime import datetime, timedelta, date
from openpyxl.utils import get_column_letter
import time
import re
import string
import copy


conversion = {'Ajax, DoorProtect v.8a_868 Mhz_DIP' : ['dpv8'],
              'Ajax, SpaceControl v8_868 Mhz_DIP' : ['sc','scv8'],
              'Ajax, DP.011.MBR.001v3_DIP' : ['dpv3', 'dp'],
              'Ajax, MotionProtect v4_А_868 Mhz_DIP' : ['mp', 'mpv4c', 'mpv4', 'mpv4a'],
              'Ajax, GlassProtect v4_868 Mhz_DIP (28.02.2020)' : ['gpv4'],
              'Ajax, ocBridge v.15_868 Mhz_DIP v1 (12.02.2020)' : ['ocbridge', 'oc','ocbridgev15'],
              'Ajax, uartBridge v.8_868 Mhz_DIP' : ['uartbridge', 'uart','uartbridgev8'],
              'Ajax, HUB.001.PWB.001v1_Dip' : ['pwbv1','hubpwbv1'],
              'Ajax, HUB-07e_868 Mhz_DIP' : ['hub07e' , 'hub'],
              'Ajax, HUB-07d_868 Mhz_DIP' : ['hub07d'],
              'Ajax, StreetSiren_v5_868 Mhz_DIP (31.03.2020)' : ['ssv6'],
              'Ajax, WallSwitch_Full_DIP' : ['ws','wsv8','wsv9'],
              'Ajax, LeaksProtect_v8_868 Mhz_DIP' : ['lp', 'lpv8'],
              'Ajax, FP_.001.MBR.000v7d_DIP' : ['fp', 'fpv7d', 'fpv7'],
              'Ajax, Fire_Protect_Top_Board_V2.1_Dip' : ['topboard', 'tp','tobbord','fptb'],
              'Ajax, FPP_.001.MBR.000v7d_DIP' : ['fpp', 'fppv7d'],
              'Ajax, CombiProtect_M_v4_A_868 Mhz_DIP (24.01.2020)' : ['cp','mpp','cpv4c','mppv4c','сp'],
              'Ajax, HS.001.MBR.001_v4_868 Mhz_DIP' : ['hs','hsv4'],
              'Ajax, Keypad V12_868 Mhz_DIP' : ['kp','kpv12','кр'],
              'Ajax, DoorProtect Plus_ v.9_868 Mhz_DIP' : ['dpp','dppv9'],
              'Ajax, TRS.002.601.000v3_Dip' : ['trs','trsv3'],
              'HUB.001.601.000v16' : ['hubplus','hubv16','hub+','hub+v16','hbv16'],
              'Ajax, MPO.001.MBR.001v5_868 Mhz_DIP' : ['mpo','mpov5','mpombr','mpombrv5'],
              'Ajax, MPO.001.PIR.001v5_DIP' : ['mpopir','mpopirv5'],
              'Ajax, MPC.001.601.000-01v8_DIP' : ['mpcpir','mpcpirv8','mpcv8'],
              'Ajax, MPC.001.603.000-01v7_868 Mhz_DIP' : ['mpc','mpcmbr', 'mpcv7','mpcmbrv7','mpcmrb'],
              'Ajax, ReX_Hub-07d_868 Mhz_DIP' : ['rex07d'],
              'Ajax, HB2.001.MBR.001v3_868 Mhz_DIP' : ['hub2v3'],
              'Ajax, CAM.001.MBR.001v7_868 Mhz_DIP' : ['mcamv7','mcam','mcammbr','camv7','cammbrv7','mcv7','мcamv7'],
              'Ajax, CAM.001.PIR.001v7_868 Mhz_DIP' : ['campirv7','campir','mcampirv7','cimpirv7'],
              'Ajax, CAM.004.MBR.001v1_868 Mhz_DIP' : ['mcamv1'],
              'Ajax, CAM.004.PIR.001v0_868 Mhz_DIP' : ['campirv0','mcampirv0','campirvo'],
              'Ajax, PBJ.003.MBR.001v5_868 Mhz_DIP' : ['pbj','button','pbjv5','batton'],
              'Ajax, HB2.001.P12.001v8_DIP' : ['p12forhub2'],
              'Ajax, HUB.001.P12.001v8_DIP' : ['p12forhub'],
              'Ajax, HB2.002.MBR.001v2a_DIP' : ['hub2plus', 'hub2plusv2a', 'hub2+','hub2+v2a'],
              'Ajax, MTR.001.MBR.000v1b_DIP_KLS' : ['mtrv1b', 'mtr'],
              'Ajax, HB2.001.P6V.001v1_DIP' : ['p6forhub2','psu6vforhub2','hub2p6v','p6fothub2','hb2p6v'],
              'Ajax, PBD.001.MBR.001v2_DIP' : ['pbd', 'doublebutton', 'dbutton','pbdv2'],
              'Ajax, SOC All v9a' : ['soc.full', 'socketfull', 'soc.fullv9','socketfullv9','soc.fullv9a','socfull','socfullv9a','soketfulv9','socall', 'socketall','sokfull'],
              'Ajax, DCO.001.MBR.001v5_DIP' : ['dco', 'dcov5','dcombr','dcombrv5'],
              'Ajax, DCO.001.PIR.001v4_DIP' : ['dcopir','dcopirv4'],
              'Ajax, KPP.001.MBR.001v12_DIP' : ['kpp', 'kppv1', 'kppv2'],
              'Ajax, KPS.001.BOT.001v4_DIP' : ['kpsbot','kpcbot', 'kpsbotv4'],
              'Ajax, MCO.001.PIR.001v3_DIP' : ['mcopirv3'],
              'Ajax, MCO.001.MBR.001v5_DIP' : ['mcov5', 'mcombrv5'],
              'Ajax, MCO.002.MBR.001v0_DIP' : ['mcov0', 'mcombrv0', 'mcophod', 'mcov0phod','mcombrv0'],
              'Ajax, MCO.001.OVB.001v6_DIP' : ['ovbv6', 'mcoovb','mcoovbv6','obvv6'],
              'Ajax, SS.002.MBR.001v5_DIP' : ['ssv5', 'ss'],
              'Ajax, HB2.001.MBR.001v3 (RX2-8XX)_DIP' : ['rex2v4', 'rex2'],
              'Ajax, VHF.001.MBR.001v5_DIP' : ['vhfbridge', 'vhf','vhfv5', 'vhfmbrv5', 'vhfbridgev5','vhfbridg'],
              'Ajax, HB2.003.MBR.001v1_DIP' : ['hub24gv1', 'hub24g','hb2003v1','hub24g','hb24g'],
              'Ajax, HB2.001.PWB.001v4_DIP' : ['pwbv4', 'hubpwbv4', 'hub2pwbv4'],
              'Ajax, DPF.001.MBR.001_v4 DIP' : ['dpf','dpfv4'],
              'Ajax, HSF.001.MBR.001v3_868 Mhz_DIP' : ['hsf', 'hsfv3'],
              'Ajax, GPF.001.MBR.001v15_868 Mhz_DIP' : ['gpf', 'gpfv15'],
              'Ajax, KPF.001.MBR.001v4_868 Mhz_DIP (установочная партия)' : ['kpf', 'kpfv4'],
              'Ajax, MPF.001.MBR.001v13[MPP]_DIP' : ['mppfv13', 'mppf'],
              'Ajax, MPF.001.MBR.001v13_DIP' : ['mpfv13', 'mpf'],
              'Ajax, GP.001.MBR.001v10_868 Mhz_DIP (установочная партия)' : ['gpv10', 'gp','gp10'],
              'Ajax, HB2.001.P24.001v1_DIP (установочная партия)' : ['p24forhub2','р24forhub2','hub2p24','hb2p24','psup24forhub2','hb2p24v1'],
              'Ajax, HB2.001.MBR.001v4_868 Mhz_DIP (установочная партия)' : ['hub2v4'],
              'Ajax, SOG.001.REL.001v3_DIP (установочная партия)' : ['sog.relv3'],
              'Ajax, Socket type G_Full_DIP' : ['sogv5','sockettypegv5','soctypegv5','sockettypeg','soctypeg','soc.gv5','soc.typeg'],
              'Ajax, MP.002.MBR.001v0_868 Mhz_DIP (установочная партия)' : ['mpv0','мpv0'],
              'Ajax, HUB-07e [REX-07e-3P-8XX]_DIP' : ['rex07e','rex','rex07e3p','rex3pin','rex07e3pin','rex3p','hb07e3prex'],
              'Ajax, HUB.001.PWB.001v7_Dip (установочная партия)' : ['pwbv7','hubpwbv7','hbpwbv7','pwdv7'],
              'Ajax, HUB-07e(Hub-07e-3P-8XX)_DIP' : ['hub07e','hub3pin','hub07e3pin','hub07e3p','hub3p'],
              'Ajax, GP.001.MBR.001v10 (9XX)_DIP' : ['gpv10915','gp915','gp(915)'],
              'Ajax, CAM.001.PIR.001v7 (9XX)_DIP' : ['campirv7915','campir915','campirv7(915)','mcampirv7915'],
              'Ajax, CAM.004.PIR.001v0_(9XX)_DIP (установочная партия)' : ['campirv0915','campirv0(915)'],
              'Ajax, SpaceControl v8 (9XX)_DIP' : ['sc915','scv8915','sc(915)'],
              'Ajax, DP.001.MBR.001v9 (9XX)_DIP' : ['dp915','dpv9','dpv9915'],
              'Ajax, HB2.001.MBR.001v4 (9XX)_DIP' : ['hub2v4915'],
              'Ajax, HS.001.MBR.001v5 (9XX)_DIP' : ['hs915', 'hsv5915'],
              'Ajax, HB2.002.MBR.001v2a_(9XX/ECG)_DIP' : ['hub2plus915ecg','hub2+915ecg','hub2+ecg','hub2plusecg','hub2+915'],
              'Ajax, HB2.002.MBR.001v2a_(9XX/AFA)_DIP' : ['hub2plus915afa','hub2plusafa','hub2+915afa','нub2plus915afa'],
              'Ajax, HB2.002.MBR.001v2a_(9XX/AUX)_DIP' : ['hub2plus915aux','hub2plusaux','hub2+915aux'],
              'Ajax, HB2.002.MBR.001v2a (9XX_AUX_PSG)_DIP' : ['hub2plus915aux-psg','hub2plus915auxpsg','hub2plusauxpsg'],
              'Ajax, Keypad V12 (9XX)_DIP' : ['kp915','kpv12915','kpv12(915)'],
              'Ajax, CP_M_v4c_Motion (9XX)_DIP' : ['mp915','mpv4c915','mpv4915','mp4c915'],
              'Ajax, CP_M_v4c_Combi_Motion Plus (9XX)_DIP' : ['cp915','cpv4c915','mpp915','mppv4c915'],
              'Ajax, DPP.001.MBR.001v9 (9XX)_DIP' : ['dpp915','dppv9915'],
              'Ajax, PBJ.903.MBR.001v0 (9XX)_DIP' : ['pbj915','pbjv0915','button915','pbjvo915'],
              'Ajax, WSD.001.601.000v11 (9XX)_Dip' : ['wsd915','wsdv11915'],
              'Ajax, WSD.001.604.000v11_Dip' : ['wsd','wsdv11'],
              'Ajax, LeaksProtect_v8 (9XX)_DIP' : ['lp915','lpv8915'],
              'Ajax, SS.002.MBR.001v5 (9XX)_DIP' : ['ssv5915','ss915'],
              'Ajax, MPC.001.603.000-01v7 (9XX)_DIP' : ['mpc915','mpcv7915'],
              'Ajax, MPO.001.MBR.001v5 (9XX)_DIP' : ['mpo915','mpov5915','mpombr915'],
              'Ajax, TRS.002.601.000v3 (9XX)_Dip' : ['trs915','trsv3915'],
              'Ajax, MTR.001.MBR.000v1b (9XX)_DIP' : ['mtr915','mtrv1b915'],
              'Ajax, HUB-07e [REX-07e-3P-9XX]_DIP' : ['rex07e915','rex915'],
              'Ajax, DCO.001.MBR.001v5 (9XX)_DIP' : ['dco915','dcov5915','dcombr915'],
              'Ajax, KPP.001.MBR.001v12 (9XX)_DIP' : ['kpp915','kppv12915'],
              'Ajax, WallSwitch_Full_915_DIP' : ['ws915','wsv8915','wsv9915'],
              'Ajax, SOC.001.REL.001v9_Relay Board_DIP' : 'soc601-602',
              'Ajax, SOC.001.PLG.001v9a_Plug Board_DIP' : 'soc603',
              'Ajax, SOC.001.SOC.001v9a_Socket Board_DIP' : ['soc604'],
              'Ajax, HB2.001.MBR.001v4 (RX2-9XX)_DIP' : ['rex2v4915'],
              'Ajax, SOC All v9a (9XX)' : ['soc.full915','soc.fullv9a915','socket915','socfull915'],
              'Ajax, MCO.001.MBR.001v5 (9XX)_DIP' : ['mcov5915','mco915','mcombr915','mcombrv5915'],
              'Ajax, MCO.001.FLB.001v4_плата с компонентами' : ['flb','mcoflb','flbv4','mcoflbv4'],
              'Ajax, MCO.002.MBR.001v0 (9XX)_DIP (установочная партия)' : ['mcov0915','mcombrv0915'],
              'Ajax, HB2.003.MBR.001v1_(9XX/AUX)_DIP' : ['hub24gv1915','hub24gv1915aux'],
              'Модуль камери ЧОРНІ RH-AA508A03D11-46 BF3005 1/4" 640x480 VGA YXF' : ['модулькамеричорні'],
              'Модуль камери ЖОВТІ YXF-HDF7740-A46-H95-0Q' : ['модулькамерижовті'],
              'RFID модуль MIFARE DesFire EV1 2K, чип+антенна, Rантенны=26мм CKT' : ['tag'],
              'Ajax, KPS.001.TOP.001v6_плата с компонентами' : ['kpstop','kpc','kpstopv6','kpsv6','kpstopv7','kpsv7','kpcv6','kpcv7'],
              'Ajax, CFM.001.MBR.001v2_плата с компонентами' : ['cfmv2'],
              'Ajax, CFM.001.MBR.001v1_плата с компонентами' : ['cfmv1'],
              'Ajax, MCO.001.OVB.001v4_DIP' : ['ovbv4','mcoovbv4'],
              'Ajax, HB2.005.MBR.001v0_[HB2-8XX]_DIP (установочная партия)' : ['hub2v0','hub2','hb2v0','hb2','hb2vo'],
              'Ajax, HB2.003.MBR.001v1_[8XX-ECG-EA]_DIP' : ['hub24gv1звиноснимиантенами','hb2.003v1ea','hb2.003v1ea','hub24gv1звин.ант.'],
              'Ajax, HB2.006.MBR.001v0_[8XX-ECG]' : ['hub24gv0','hb24gvo','hb24gvo','hub24gv0'],
              'Ajax, LSW.001.MBR.001v6[1gang]_плата с компонентами' : ['lsw1gang','lswmbr1gang','lsw1-gang','lswmbr1-gang', 'lsw1g','lswmbr1g'],
              'Ajax, LSW.001.PWR.001v9[1gang]_DIP' : ['pwr1gang','lswpwr1gang','pwr1-gang','lswpwr1-gang','pwr1g','lswpwr1g'],
              'Ajax, LSW.001.MBR.001v6[2gang]_плата с компонентами' : ['lsw2gang','lswmbr2gang','lsw2-gang','lswmbr2-gang','lsw2g','lswmbr2g'],
              'Ajax, LSW.001.PWR.001v9[2gang]_DIP' : ['pwr2gang','lswpwr2gang','pwr2-gang','lswpwr2-gang','pwr2g','lswpwr2g'],
              'Ajax, LSW.001.MBR.001v6[2way]_плата с компонентами' : ['lsw2way','lswmbr2way','lsw2-way','lswmbr2-way','lsw2w','lswmbr2w'],
              'Ajax, LSW.001.PWR.001v9[2way]_DIP' : ['pwr2way','lswpwr2way','pwr2-way','lswpwr2-way','pwr2w','lswpwr2w'],
              'Ajax, HBF.001.MBR.001v28_DIP (установочная партия)' : ['hbfv28','hbf'],
              'Ajax, HB2.005.MBR.001v0_[RX2-8XX]_DIP (установочная партия)' : ['rex2v0','rex2','rex2v0','rex2v0'],
              'Ajax, LQC.001.MBR.001v10' : ['lqc','lqcv10','lq','lqv10'],
              'Ajax, LQC.001.LED.001v4' : ['lqcled','ledv4','lqcledv4','lqled'],
              'Ajax, LQC.001.MBR.001v10 (9XX)' : ['lqc915','lqcv10915','lq915','lqv10915','lqcmbrv10915'],
              'Ajax, MTF.001.MBR.001v2_DIP (установочная партия)' : ['mtfv2','mtf'],
              'Ajax, HB2.006.MBR.001v0_(9XX/AUX)_DIP    (4G)' : ['hub24gv0915aux'],
              'Ajax, HB2.005.MBR.001v0_[RX2-9XX]_DIP (установочная партия)' : ['rex2v0915','rex2915','rex2v066a915'],
              'Ajax, HB2.005.MBR.001v0_[HB2-9XX]_DIP (установочная партия)' : ['hub2v0915','hub2915'],
              'Ajax, SSF.001.MBR.001v9_DIP (установочная партия)' : ['ssf','ssfv9'],
              'Ajax, EM4.002.MBR.001v4_[FP_8ХХ]_DIP2' : ['fp2rb','fp2rbhs','fp2rbhs', 'fp2rbheatsmoke','fp2rbsh','em4rbhs'],
              'Ajax, EM4.002.MBR.001v4_[FPP_8ХХ]_DIP2' : ['fpp2rb','fp2rbhsco','fp2rbhsco'],
              'Ajax, EM4.002.MBR.001v4_[FP2_SB_8ХХ]_DIP2' : ['fp2sb','fp2sbhs','fpsrbhs', 'fp2sbheatsmoke'],
              'Ajax, EM4.002.MBR.001v4_[FPP2_SB_8ХХ]_DIP2' : ['fpp2sb','fp2sbhsco','fp2sbhsco'],
              'EM4.002.MBR.001v4_[FP_8/9ХХ]  (915)' : ['fp2rb915','fp2rbhs915','fp2rbhs915', 'fp2rbheatsmoke915'],
              'EM4.002.MBR.001v4_[FPP_8/9ХХ]  (915)' : ['fpp2rb915','fp2rbhsco915','fp2rbhsco915'],
              'EM4.002.MBR.001v4_[FP2_SB_8/9ХХ]  (915)' : ['fp2sb915','fp2sbhsco915','fp2sbhsco915'],
              'EM4.002.MBR.001v4_[FPP2_SB_8/9ХХ]  (915)' : ['fpp2sb915', 'fp2sbhsco915','fp2sbhsco915'],
              'Ajax, HB2.005.MBR.001v0_[RX2-8XX]_DIP (установочная партия)' : ['rex2v066a','rex2v066'],
              'Ajax, HB2.005.MBR.001v0[HB2-8XX-66A] (установочная партия)' : ['hub2v066a','hub2vo66a','hub2rfm-66a'],
              'Ajax, HB2.006.MBR.001v0[8XX-ECG-66A] (установочная партия)' : ['hub24gv066a','hub24gv066','hub24gv0rfm66-a','hub24grfm-66a'],
              'Ajax, Socket type G_Full_(9SA)_DIP' : ['sogv3','sockettypegv3','soctypegv3','soc.gv3'],
              'Ajax, HB2.008.MBR.001v1[8XX_ECG]_DIP' : ['hub2plusv1','hub2+v1','hb2plusv1'],
              'Ajax, Socket Plus (type G)_Full_DIP' : ['sgp','socketplustypeg','socfullplustypeg','socket+g'],
              'Ajax, HB2.005.MBR.001v0[HB2-9XX-66A] (установочная партия)' : ['hub2v066a915','hub2v066а915'],
              'Ajax, CAM.004.MBR.001v3_(8XX)_DIP' : ['camphod','mcamphod','mcamphodv3','camphodv3','mcamhpodv3','camv3phod'],
              'Ajax, MPF.002.MBR.001v1[MPP]_DIP' : ['mppfv1','mpfv1'],
              'Ajax, TBF.001.MBR.001v2_DIP' : ['tbf', 'tbfv2'],
              'Ajax, MCF.002.MBR.001v2_DIP' : ['mcfv2','mcfmbrv2'],
              'Ajax, MCF.002.PIR.001v2_DIP' : ['mcfpirv2'],
              'Ajax, MCF.002.MBR.001v2_[MCF_B]_DIP' : ['mcfphodv2','mcfphod'],
              'Ajax, MCO.001.OVB.001v8_DIP' : ['ovbv8', 'mcoovbv8','mcov8','ovb'],
              'Ajax, MCO.001.MBR.001v11 (8XX)_DIP' : ['mcov11', 'mcombrv11', 'mco', 'mcombr','мсov11'],
              'Ajax, MCO.001.PIR.001v5_DIP' : ['mcopirv5','mcopirv5','mcopir'],
              'Ajax, HB2.006.MBR.001v0[9XX-AUX-66A]_DIP' : ['hub24gaux', 'hub2aux4g','hub24gaux66a','hub24gv0aux','hub24gv0915'],
              'Ajax, MCO.002.MBR.001v1 (8XX)_DIP' : ['mcov1','mcophodv1','mcophod','mсоv1','mcоv1','mcombrv1'],
              'Ajax, HBF.002.MBR.001v2_DIP (установочная партия)' : ['hbfv2','hbf4g','hbfv24g','hbf4gv2'],
              'EM4.002.MBR.001v4_[FP_8/9ХХ] (915)' : ['fp2rb915','fp2rbhs915','fp2rbh/s915','fp2rbheatsmoke915','em4rbhs915'],
              'HB2.008.MBR.001v1_[9XX-AUX]' : ['hub2plusaux','hub2plusv1aux','hb2plusv1aux','hub2+v1aux','hub2+aux','hub2plus915aux','hub2plusaux915','hub2plus008v1915aux','hub2plusv1915aux','hub2+v1915aux','hub2plusv1aux915'],
              'HB2.008.MBR.001v1_[9XX-AFA]' : ['hub2plusafa','hub2plusv1afa','hb2plusv1afa','hub2+v1afa','hub2+afa','hub2plus915afa','hub2plusafa915','hub2plus008v1915afa','hub2plusv1915afa','hub2+v1915afa','hub2plusv1afa915'],
              'Ajax, EM2.001.MBR.001v4_[FP2_C_SB_9ХХ]_DIP' : ['fp2sbco915','fp2cosb915','fp2sbc915','em2sbco915','em2sbc915'],
              'Ajax, EM2.001.MBR.001v4_[FP2_С_RB_9ХХ]_DIP' : ['fp2rbco915','fp2corb915','fp2rbc915','em2rbco915','em2rbc915'],
              'Ajax, EM2.001.MBR.001v4_WH_[FP2_H_SB_9ХХ]_плата с компонентами' : ['fp2sbheat915white','fp2heatsb915wl','fp2sbh915wl','em2sbheat915white','em2sbh915wl'],
              'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_SB_9ХХ]_плата с компонентами' : ['fp2sbheat915black','fp2heatsb915bl','fp2sbh915bl','em2sbheat915black','em2sbh915bl'],
              'Ajax, EM2.001.MBR.001v4_[FP2_C_SB_8ХХ]_DIP' : ['fp2sbco','fp2cosb','fp2sbc','em2sbco','em2sbc','em2cosb','fp2minisbco'],
              'Ajax, MTF.001.MBR.001v1' : ['mtfv1','mtrfv1'],
              'Ajax, HB2.006.MBR.001v0_[9XX-ECG-66A]_DIP' : ['hub24gsa','hub24gsa915','hub24g915sa','hub24gv0esg915'],
              'Ajax, CAM.004.PIR.001v0_(9XX)_DIP' : ['campirv0915','campirphod915'],
              'Ajax, KPS.001.TOP.001v6 (9XX)_плата с компонентами' : ['kpstop915]','kpstopv6915'],
              'Ajax, HB2.003.MBR.001v1_(9XX/ECG)_DIP' : ['!'],
              'Ajax, HB2.003.MBR.001v1_(9XX/AFA)_DIP' : ['!'],
              'Ajax, DPF.001.MBR.001_v4[DPP] DIP' : ['dppf', 'dppfv4'],
              'Ajax, HB2.001.MBR.001v4 (RX2-8XX)_DIP (установочная партия)' : ['rex2v4'],
              'Ajax, HB2.006.MBR.001v0[9XX-AUX-PSG]_DIP' : ['hub24gauxpsg', 'hub24g915auxpsg','hub24gv0auxpsg'],
              'Ajax, HB2.006.MBR.001v0_[8XX-ECG-PSG]_DIP' : ['hub24gpsg','hub24gv0psg'],
              'Ajax, MCO.001.MBR.001v11 (9XX)_DIP' : ['mcov11915','mcombrv11915'],
              'Ajax, CAM.007.MBR.001v3_DIP' : ['mcamv3','camv3','mcammbrv3','cammbrv3','mcv3'],
              'Ajax, MCF.002.MBR.001v4_[MCF_G]_DIP' : ['mcfv4','mcfmbrv4','mcf','mcfmbr'],
              'Ajax, MCF.002.PIR.001v4_[PIR_B]_DIP' : ['mcfpirv4','mcfpir','mcfv4pir'],
              'Ajax, DP.012.MBR.001v1_DIP' : ['dpv1','dps-line','dps'],
              'Ajax, DPP.012.MBR.001v1_DIP' : ['dppv1','dpps-line','dpps','dpsp'],
              'Ajax, MP.002.MBR.001v0_(9XX)_DIP' : ['mpv0915','mp915v0'],
              'Ajax, HB2.006.MBR.001v0[9XX-AUX-PSG]_DIP' : ['hub24gauxpsg','hub24g915auxpsg','hub24g915psg','hub24gauxpsg915','hub24gpsgaux','hub24gpsgaux915'],
              'Ajax, HYB.HB1.MBR.001v2_868 Mhz_DIP' : ['yavir','yavir+','yavirplus','yavirv2','yavir+v2','yavirplusv2','hybv2','yyvir','hybmbrv2'],
              'Ajax, HB2.006.MBR.001v0_(9XX/AUX)_DIP (4G)' : ['hub24gaux','hub24g915aux','hub24gaux915','hub24gv0aux','hub24gv0aux915','hub24g915'],
              'Ajax, MCO.002.MBR.001v1 (9XX)_DIP' : ['mcov1915'],
              'Ajax, HUB.001.P24.001v1_DIP (установочная партия)' : ['p24forhub','р24forhub','hubp24','psup24forhub'],
              'Ajax, EM2.001.MBR.001v4_[FP2_С_RB_8ХХ]_DIP' : ['fp2rbco','fp2rbc','em2rbc','em2rbco','em2v4rbco','fpp2em2corb','fp2minirbco','fp2rbcomini'],
              'Ajax, EM2.001.MBR.001v4_WH_[FP2_H_RB_8ХХ]_плата с компонентами' : ['fp2rbheatwhite','em2rbhw','em2rbheatwhite','fp2em2rb'],
              'Ajax, EM2.001.MBR.001v4_WH_[FP2_HС_RB_8ХХ]_плата с компонентами' : ['em2rbhcowhite','fp2rbhcow','fp2rbhcowhite','fp2rbhcw','fp2rbhcwhite'],
              'Ajax, HB2.008.MBR.001v1[9XX_ECG]_DIP' : ['hub2+sa915','hub2plussa915','hub2plus915saecg','hub2+915saecg','hub2plussaecg','hub2+saecg','hub2plussaecg915','hub2+saecg915','hub2plussa915','hub2+sa915','hub2plusv1ecg915'],
              'Ajax, EM2.001.MBR.001v4_WH_[FP2_HC_SB_8ХХ]_плата с компонентами' : ['fp2heatcosbw','em2hcsbw','em2heatcosbw','fp2sbhcw','fp2sbheatcow','em2sbwcoh'],
              'Ajax, HBF.002.MBR.001v4_DIP' : ['hbf4gv4'],
              'Ajax, WTS.001.MBR.001v4 (8XX)_DIP' : ['wtsmbr','wtsmbrv4','wtsv4','wtsv4mbr','wtrmbr','wtrmbrv4','wtrv4mbr'],
              'Ajax, WTS.001.MBR.001v4 (9XX)_DIP' : ['wtsmbr915','wts915','wtsv4915','wtsmbrv4915','wts915v4','wts915mbr','wts915mbrv4','wtrmbr915','wtrmbrv4915','wtsmrb915','wts'],
              'Ajax, WTS.001.PWB.001v4 (8XX)_DIP' : ['wtspwb','wtspwbv4','wtsv4pwb','wtrpwb','wtrpwbv4','wtrv4pwb'],
              'Ajax, MPF.002.MBR.001v4_DIP' : ['mpfv4','mppfv4'],
              'Ajax, HB2.006.MBR.001v0[8XX-ECG-66A-VPS]_DIP' : ['hyb24g6vvps','hub24gvps','hb24g6vvps','hb24gvps','hub24g6v'],
              'Ajax, NVR.003.MBR.001v5_DIP' : ['nvr','nvrmbr','nvrv5','nvrmbrv5','ananas','nvrv3mbr','mvrmbr'],
              'Ajax, NVR.003.PWB.001v3_DIP' : ['nvrpwb','nvrpwbv3','pwbnvr','nvrv3pwb']
              
}

not_used = ['Ajax, CP_001.601.000v3_DIP (24.01.2020) мікрофон',
            'Ajax, CombiProtect_M_v4_A_868 Mhz_DIP (24.01.2020) мікрофон',
            'Ajax, CP_M_v4c_Combi_Motion Plus_DIP новий мкрофон',
            'Ajax, Hub_V4_holder_DIP',
            'Ajax, StreetSiren_DG DIP',
            'Ajax, StreetSiren_Buzzer DIP',
            'Ajax, StreetSiren_LED1_v3, LED2_v3_DIP старі',
            'Ajax, StreetSiren_Battery_v3_DIP  Ліва сторона (для старої версії плати)',
            'Ajax, StreetSiren_Battery_v3_DIP Права сторона (для старої версії плати)',
            'Ajax, StreetSiren_LED1_v3, LED2_v3_DIP',
            'Ajax, LeaksProtect 2P_v8_DIP',
            'Ajax, StreetSiren_Battery_v3_DIP Ліва сторона (для старої версії плати)',
            'Ajax, CP_M_v4c_Combi_Motion Plus_DIP',
            'Ajax, MotionProtectMW V2_DIP',
            'Ajax, MotionProtectMW V2 (XUK)_DIP (установочная партия)',
            'Ajax, Relay_Full_(868)_DIP',
            'Ajax, HB2.001.PWB.001v2_DIP',
            'Ajax, YVH.001.MBR.001_v3_868 Mhz_DIP',
            'Ajax, CAM.001.MBR.001v11_868 Mhz_DIP',
            'Ajax, CAM.001.PIR.001v11_868 Mhz_DIP',
            'Ajax, HB2.001.PWB.001v2_DIP',
            'Ajax, HB2.002.MBR.001v2a (8XX_ECG_PSG)_DIP',
            'Ajax, HB2.001.PWB.001v2_DIP',
            'Ajax, MTR.001.MBR.000v1b_DIP_KF (UA)',
            'Ajax, MTR.001.MBR.000v1b_DIP',
            'Ajax, KPP.001.ANT.001v9_DIP',
            'Ajax, KPS.001.BUZ.001v3_DIP',
            'Ajax, MCO.001.MBR.001v7_DIP (установочная партия)',
            'Ajax, MCO.001.MBR.001v8_DIP',
            'Ajax, StreetSiren_LD1, LD2_DIP',
            'Ajax, SS.002.BT1.001v1_DIP',
            'Ajax, SS.002.BT2.001v2_DIP',
            'Ajax, HB2.001.PWB.001v2_DIP',
            'Ajax, HBF.001.MBR.001v26_DIP (установочная партия)',
            'Ajax, HSF.001.CON.001v3_868 Mhz_DIP',
            'Ajax, KPF.001.CON.001v2_868 Mhz_DIP',
            'Ajax, CPF.001.GP.001v2_DIP',
            'Ajax, MPF.001.CON.001v13_DIP',
            'Ajax, MPF.001.CON.001v13_DIP',
            'Ajax, MPF.001.CON.001v13_DIP',
            'Ajax, MotionProtectMW V2_DIP',
            'Ajax, SOG.001.PWR.001v3_DIP (установочная партия)',
            'Ajax, MTR.001.MBR.000v3_DIP (установочная партия)',
            'Ajax, HUB.001.PWB.001v5_Dip (установочная партия)',
            'Ajax, SC.011.MBR.001v1_868 Mhz_DIP (установочная партия)',
            'Ajax, NVR.003.MBR.001v3_DIP (установочная партия)',
            'Ajax, NVR.003.PWB.001v0_DIP (установочная партия)',
            'Ajax, HBL.001.MBR.001v4_DIP (установочная партия)',
            'Ajax, LSW.001.PWR.001v8[1gang]_DIP',
            'Ajax, LSW.001.PWR.001v8[2gang]_DIP',
            'Ajax, LSW.001.PWR.001v8[3way]_DIP',
            'Ajax, HB2.001.PWB.001v4_DIP (установочная партия)',
            'Ajax, HUB-07e (Hub-07e-3P-9XX) Mhz_DIP',
            'Ajax, HUB.001.PWB.001v5 (9XX)_Dip',
            'Ajax, HB2.001.MBR.001v3 (9XX)_DIP',
            'Ajax, HB2.001.PWB.001v2 (9XX)_DIP',
            'Ajax, CP_001.601.000v3_DIP (24.01.2020)',
            'Ajax, MotionProtectMW V2_DIP',
            'Ajax, MotionProtectMW V2 (XNA)_DIP (установочная партия)',
            'Ajax, Relay_Full_(915)_DIP',
            'Ajax, StreetSiren_LD1, LD2_DIP',
            'Ajax, SS.002.BT1.001_DIP',
            'Ajax, SS.002.BT2.001_DIP',
            'Ajax, StreetSiren_LD1, LD2_AU_DIP',
            'Ajax, HUB.001.PWB.001v5 (9XX)_Dip',
            'Ajax, KPP.001.ANT.001v9_DIP',
            'Ajax, KPS.001.BUZ.001v3_DIP',
            'Ajax, WS.001.PWB.001v8_DIP',
            'Ajax, WS.001.RFM.001v8 (9XX)_DIP',
            'Ajax, FP_.001.MBR.000v7d[9XX]_DIP',
            'Ajax, FireProtect_CO_Board_v6_ Dip',
            'Ajax, FPP_.001.MBR.000v7d[9XX]_DIP',
            'Ajax, FireProtect_CO_Board_v6_ Dip',
            'Ajax, SOC.001.MBR.001v9_Radio Board_плата с компонентами',
            'Ajax, HB2.001.MBR.001v3 (RX2-9XX)_DIP',
            'Ajax, HB2.001.PWB.001v2_DIP',
            'Ajax, HB2.001.PWB.001v2 (9XX)_DIP',
            'Ajax, SS.002.BT2.001v3_DIP',
            'Ajax, UTB.001.601.000v5_DIP',
            'Ajax, MCO.001.MBR.001v10_DIP (установочная партия)',
            'Ajax, WS.001.RFM.001v8_DIP',
            'Ajax, HB2.001.MBR.001v4 [RX2-9XX]_DIP',
            'HB2.008.MBR.001v0_[8XX-ECG]',
            'HB2.008.MBR.001v0_9XX',
            'Ajax, SS.002.MBR.001v5[9XX-66A]_DIP (установочная партия)',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_RB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_RB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_HС_RB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_HС_RB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_SB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_SB_8ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_HC_SB_8ХХ]_плата с компонентами',
            'Ajax, EM4.002.MBR.001v4_WH_[FP2_HS_RB_8ХХ]_DIP2',
            'Ajax, EM4.002.MBR.001v4_BL_[FP2_HS_RB_8ХХ]_DIP2',
            'Ajax, EM4.002.MBR.001v4_BL_[FP2_HSC_RB_8ХХ]_DIP2',
            'Ajax, EM4.002.MBR.001v4_WH_[FP2_HSC_RB_8ХХ]_DIP2',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_HС_RB_9ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_WH_[FP2_HС_RB_9ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_H_RB_9ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_WH_[FP2_H_RB_9ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_BL_[FP2_HC_SB_9ХХ]_плата с компонентами',
            'Ajax, EM2.001.MBR.001v4_WH_[FP2_HC_SB_9ХХ]_плата с компонентами',
            'EM4.002.MBR.001v4_[FPP_8/9ХХ] (915)',
            'EM4.002.MBR.001v4_[FP2_SB_8/9ХХ] (915)',
            'EM4.002.MBR.001v4_[FPP2_SB_8/9ХХ] (915)',
            'Ajax, SOC.001.PLG.000v10_DIP (установочная партия)'
            
]



def input_date(): #визначаємо дату перевірки
    days_of_week = {'Sunday' : 'нд', 'Monday' : 'пн', 'Tuesday' : 'вт', 'Wednesday' : 'ср', 'Thursday' : 'чт', 'Friday' : "пт", 'Saturday' : 'сб'}
    choose_date = datetime.strptime(input('Введіть дату для перевірки у форматі Рік.місяць.день (Напр. 2023.03.18): ' ), "%Y.%m.%d").date()
    date_for_inspection =(choose_date.strftime("%Y.%m.%d.%A"))
    for key,value in days_of_week.items():
        if key in date_for_inspection:
            day_of_week = value
    date_for_inspection = f"{choose_date.strftime('%Y.%m.%d')}.{day_of_week}"
    print(f"День перевірки - {date_for_inspection}")
    print('-----------------------------------------------------------------------')
    print('\n')
    choose_date = str(choose_date)
    return choose_date, date_for_inspection







class List_of_devices_inputs(): #отримуєм список девайсів з телеграму

    
    def __init__(self, date, correct_labels, csv_file):
        self.date = date
        self.correct_labels = correct_labels
        self.csv_file = csv_file

    def read_csv_from_telegram(self): #по рядкам зчитуєм дані з csv та прибираєм зайві символи \n 
        self.temporary_list = []
        #with open("chats.csv", "r", errors="ignore") as f:
        with open(self.csv_file, "r", errors="ignore") as f:    
            for line in f.readlines():
                self.temporary_list.append(line.strip('\n'))
        return self.temporary_list #повертаєм список з строками
        

    def strip_apostrophe(self): #прибираєм зайві символи "
        self.temporary_list_2 = []
        for i in range(len(self.temporary_list)):
            if choose_date in self.temporary_list[i]:        
                k = i
                while True:
                    k+= 1
                    if '+00:00' in self.temporary_list[k]:
                        break
                    else:
                        b = self.temporary_list[k].strip('"')
                        b = b.split()
                        b = ''.join(b)
                        self.temporary_list_2.append(b)
        return self.temporary_list_2       
        

    def separate_by_symbol (self): #розділяємо строки по "-"
        self.temporary_list_3 = []
        for i in self.temporary_list_2:
            if '-' in i:
                if any(map(str.isdigit, i)):
                    self.temporary_list_3.append([x.strip().lower() for x in i.rpartition('-')])
        return self.temporary_list_3
        

    def sort_not_need_info (self): #Відсортовуємо лишнє і залишаємо тільки девайси, перенесення в словник
        self.temporary_incomes = {}
        re_digits = re.compile(r"\b\d+\b") #шаблон для пошуку цифр у рядку
        #chars = [')', '(', '_','/'] # найчастіші помилки при ідентифікації девайсу
        chars = ['fai','prod',')', '(', '_','/','.DS_Store'] # найчастіші помилки при ідентифікації девайсу
        for i in range(len(self.temporary_list_3)):
            #res = self.temporary_list_3[i][0].translate(str.maketrans('','',''.join(chars))) #прибираєм найчастіші помилки при ідентифікації девайсу
            only_numbers = re_digits.findall(self.temporary_list_3[i][2]) #відсіюєм лишні символи після '-' і залишаєм лише числа
            res = self.temporary_list_3[i][0]
            for k in chars:
                res = str.replace(res,k, "") #прибираєм найчастіші помилки при ідентифікації девайсу
                
            try:
                if self.temporary_list_3[i][1] != '' and res not in self.temporary_incomes:
                    self.temporary_incomes[res] = int(only_numbers[0])
                elif res in self.temporary_incomes:
                    try:
                        self.temporary_incomes[res] = int(self.temporary_incomes[res]) + int(only_numbers[0])
                    except ValueError:
                        pass
            except IndexError as e:
                print(e)
                pass
            
      
        return self.temporary_incomes
        


    def manufacturing_type(self): # Заміняємо назви девайсів на повні, офіційні
        self.temporary_incomes_2 = {}
        self.incorrect_labels = set()
        for incomes_key, incomes_value in self.temporary_incomes.items():
            r = 0
            for key, value in self.correct_labels.items():
                if incomes_key in value: 
                    r += 1
                    if self.temporary_incomes_2.get(key.lower(), False) is False:  # если в словаре нет такого ключа
                        self.temporary_incomes_2[key.lower()] = [incomes_value]  # добавляем его и помещаем туда список с одним значением
                    else:  # если такой ключ уже есть
                        self.temporary_incomes_2[key.lower()].append(incomes_value)  # добавляем значение в конец списка
                    
            if r == 0:
                self.incorrect_labels.add(incomes_key)

        for key, value in self.temporary_incomes_2.items():
            self.temporary_incomes_2[key] = sum(value)

            

        return self.temporary_incomes_2

    
    def print_incorrect_labels (self): #вивести девайси, що написані неправильно
        with open('Неправильна ідентифікація.txt','a+') as f:
            line = f.readline()
            if len(self.incorrect_labels) > 0:
                f.write ( self.date + ':' + '\n' )
                print('Неправильна ідентифікація: ')
                for label in self.incorrect_labels:
                    print(label)
                    f.write(label + '\n')
                print('-----------------------------------------------------------------------')
                print('\n')
            else:
                print('Вся ідентифікація правильна ')
                print('-----------------------------------------------------------------------')
                print('\n')


                

    def print_list(self): #презентуємо наші надходження
        print('-----------------------------------------------------------------------')
        print (self.temporary_incomes_2) 
        print('-----------------------------------------------------------------------')






class Parse_google_sheet():
    def __init__(self, date_for_inspection):
        self.date_for_inspection = date_for_inspection

    def  connect_sheet(self): #підключаємось до таблиці "New Надходження Лак/ДІп/Склад"
        gs = gspread.service_account(filename='token.json')  # підключаєм файл з ключами
        sh = gs.open_by_key('1yusUPReWVPk4fr2jMtmbcxvMBoLyoJklp-B_YwAN-AM')  # підключаем таблицю по ID
        worksheet2 = sh.get_worksheet(1) # отримуєм перший лист
        self.worksheet = sh.get_worksheet(2)  # отримуєм другий лист
        return self.worksheet

    def seek_sheet(self): #шукаємо потрібні колонки з таблиці
        first_cell = self.worksheet.find(date_for_inspection) #пошук необхідної першої ячейки по даті
        try:
            self.first_column = get_column_letter(first_cell.col) #витягуєм назву колонки цієї ячейки
        except AttributeError:
            print ('Вкладку "New Надходження Лак/ДІп/Склад " перемістили або змінено часовий пояс')
        last_cell = int(first_cell.col) + 3 #Пошук останньої ячейки
        self.last_column = get_column_letter(last_cell) #витягуєм назву колонки цієї ячейки
        return self.first_column, self.last_column

    def parse_sheet(self): #збираємо дані з визначених колонок
        self.worksheet_result = {}
        new_devices = []
        for i in range(390): #390 рядків у таблиці на даний момент
            res1 = self.worksheet.get(f"A{i+2}") #зберігаємо назви девайсів з першої колонки таблиці
            if len(res1) > 0: #щоб не витрачати час на парсинг пустих рядків          
            #time.sleep(0.5)
                res2 = self.worksheet.get(f"{self.first_column}{i+2}:{self.last_column}{i+2}") #зберігаємо суми, що внесені в заданий діпазон
                try:
                    if res1[0][0] not in conversion and res1[0][0] not in not_used:
                        new_devices.append(res1[0][0])
                except IndexError:
                    pass
                sum_sheets = 0 #лічильник суми надходжень одного девайсу
                #print(res1)
                if len(res2) > 0: #записуєм суму надходжень одного девайсу
                    for k in res2[0]:
                        if k.isdigit():
                            sum_sheets += int(k)
                    self.worksheet_result[res1[0][0].lower().strip()] = sum_sheets #Зберігаємо дані у формі словника
            print(f'{i+1} ',end='')
        if len(new_devices) > 0:
            for i in new_devices:
                print(f'{i} - нова позиція, потрібно додати до проги')

            print('------------------------------------------------------------------------------------')
        return self.worksheet_result





connected_positions = {'Hub2' : ['Ajax, HB2.001.MBR.001v3_868 Mhz_DIP',
                                 'Ajax, HB2.001.MBR.001v4_868 Mhz_DIP (установочная партия)',
                                 'Ajax, HB2.005.MBR.001v0_[HB2-8XX]_DIP (установочная партия)',
                                 'Ajax, HB2.005.MBR.001v0[HB2-8XX-66A] (установочная партия)'],
                       'Hub2 4G' : ['ajax, hb2.006.mbr.001v0_[8xx-ecg]',
                                    'ajax, hb2.006.mbr.001v0[8xx-ecg-66a] (установочная партия)',
                                    'Ajax, HB2.003.MBR.001v1_DIP'],
                       'Hub2 Plus' : ['ajax, hb2.008.mbr.001v1[8xx_ecg]_dip',
                                      'ajax, hb2.002.mbr.001v2a_dip'],
                       'MP' : ['Ajax, MotionProtect v4_А_868 Mhz_DIP',
                               'Ajax, MP.002.MBR.001v0_868 Mhz_DIP (установочная партия)'],
                       'GP' : ['Ajax, GP.001.MBR.001v10_868 Mhz_DIP (установочная партия)',
                               'Ajax, GlassProtect v4_868 Mhz_DIP (28.02.2020)'],
                       'Hub' : ['Ajax, HUB-07e_868 Mhz_DIP',
                                'Ajax, HUB-07e(Hub-07e-3P-8XX)_DIP',
                                'Ajax, HUB-07d_868 Mhz_DIP'],
                       'FireProtect 2 RB (Heat/Smoke)' : ['Ajax, EM4.002.MBR.001v4_[FP_8ХХ]_DIP2',
                                                          'Ajax, EM4.002.MBR.001v4_WH_[FP2_HS_RB_8ХХ]_DIP2',
                                                          'Ajax, EM4.002.MBR.001v4_BL_[FP2_HS_RB_8ХХ]_DIP2'],
                       'FireProtect 2 RB (Heat/Smoke/CO)' : ['Ajax, EM4.002.MBR.001v4_[FPP_8ХХ]_DIP2',
                                                             'Ajax, EM4.002.MBR.001v4_BL_[FP2_HSC_RB_8ХХ]_DIP2',
                                                             'Ajax, EM4.002.MBR.001v4_WH_[FP2_HSC_RB_8ХХ]_DIP2'],
                       'Rex' : ['Ajax, ReX_Hub-07d_868 Mhz_DIP',
                                'Ajax, HUB-07e [REX-07e-3P-8XX]_DIP'],
                       'Rex 2' : ['Ajax, HB2.001.MBR.001v3 (RX2-8XX)_DIP',
                                  'Ajax, HB2.005.MBR.001v0_[RX2-8XX]_DIP (установочная партия)',
                                  'Ajax, HB2.005.MBR.001v0_[RX2-8XX]_DIP (установочная партия)',
                                  'Ajax, HB2.001.MBR.001v4 (RX2-8XX)_DIP (установочная партия)'],
                       'Hub Hybrid 2G' : ['Ajax, HBF.001.MBR.001v28_DIP (установочная партия)',
                                          'Ajax, HBF.001.MBR.001v26_DIP (установочная партия)'],
                       'CombiProtect Fibra' : ['Ajax, MPF.001.MBR.001v13[MPP]_DIP',
                                               'Ajax, MPF.002.MBR.001v1[MPP]_DIP'],
                       'Hub 2 9XX' : ['Ajax, HB2.001.MBR.001v4 (9XX)_DIP',
                                      'Ajax, HB2.005.MBR.001v0_[HB2-9XX]_DIP (установочная партия)',
                                      'Ajax, HB2.005.MBR.001v0[HB2-9XX-66A] (установочная партия)'],
                       'Hub2 Plus 9XX' : ['Ajax, HB2.002.MBR.001v2a_(9XX/ECG)_DIP',
                                          'Ajax, HB2.008.MBR.001v1[9XX_ECG]_DIP',
                                          'Ajax, HB2.002.MBR.001v2a_(9XX/AFA)_DIP',
                                          'Ajax, HB2.002.MBR.001v2a_(9XX/AUX)_DIP'],
                       'MP 915' : ['Ajax, CP_M_v4c_Motion (9XX)_DIP',
                                   'Ajax, MP.002.MBR.001v0_(9XX)_DIP'],
                       'Street Siren 9XX' : ['Ajax, SS.002.MBR.001v5 (9XX)_DIP',
                                             'Ajax, SS.002.MBR.001v5[9XX-66A]_DIP (установочная партия)'],
                       'Rex 2 9XX' : ['Ajax, HB2.001.MBR.001v4 (RX2-9XX)_DIP',
                                      'Ajax, HB2.001.MBR.001v4 [RX2-9XX]_DIP',
                                      'Ajax, HB2.005.MBR.001v0_[RX2-9XX]_DIP (установочная партия)'],
                       'Hub 2 4G NA/AUX' : ['Ajax, HB2.003.MBR.001v1_(9XX/AUX)_DIP',
                                         'Ajax, HB2.006.MBR.001v0_(9XX/AUX)_DIP (4G)',
                                         'Ajax, HB2.006.MBR.001v0[9XX-AUX-66A]_DIP']
                       
}




def duplicate(worksheet_result, finish_incomes):
    worksheet_result_new = {}
    finish_incomes_new = {}
    duplicates_in_list = []
    for key,value in connected_positions.items():
        for i in value:
            duplicates_in_list.append(i.lower())

    for key,value in connected_positions.items():
        for i in range(len(value)):
            connected_positions[key][i] = value[i].lower()


    
    for key_con, value_con in connected_positions.items():
        for key,value in finish_incomes.items():
            if key_con in finish_incomes_new and key.lower() in value_con:
                finish_incomes_new[key_con] = int(finish_incomes_new[key_con]) + int(value)
            elif key.lower() in value_con and key_con not in finish_incomes_new:
                finish_incomes_new[key_con] = value
            elif key.lower() not in duplicates_in_list:
                finish_incomes_new[key] = value
                    

    for key_con, value_con in connected_positions.items():
        for key,value in worksheet_result.items():
            if key_con in worksheet_result_new and key.lower() in value_con:
                worksheet_result_new[key_con] = int(worksheet_result_new[key_con]) + int(value)
            elif key.lower() in value_con and key_con not in worksheet_result_new:
                worksheet_result_new[key_con] = value
            elif key.lower() not in duplicates_in_list:
                worksheet_result_new[key] = value
                    
    finish_incomes = copy.deepcopy(finish_incomes_new)
    worksheet_result = copy.deepcopy(worksheet_result_new)

    return worksheet_result, finish_incomes           



            
            

def finall_check_inputs(worksheet_result, finish_incomes): #фінальна перевірка
    k = 0
    for key, value in worksheet_result.items():
        if key in finish_incomes:
            if int(value) != int(finish_incomes[key]): 
                print( '\n' + f'Неправильно внесено - {key}. Повинно бути {finish_incomes[key]}, а внесено {value}')
                k +=1 
        elif key not in finish_incomes and int(value) != 0:
            print('\n' + f"Внесено не туди. В рядку {key} повинно бути пусто, а внесено {value}")
            k +=1
        elif k == 0 :
            print ('\n' + "Обліковець не лажав, але наступного разу обов'язково напартачить")




choose_date, date_for_inspection = input_date()
start_time = datetime.now()
incomes_telegram = List_of_devices_inputs(choose_date, conversion, "chats.csv" )
incomes_telegram.read_csv_from_telegram()
incomes_telegram.strip_apostrophe()
incomes_telegram.separate_by_symbol()
incomes_telegram.sort_not_need_info()
incomes_telegram.manufacturing_type()
incomes_telegram.print_incorrect_labels()
incomes_telegram.print_list()
finish_incomes= incomes_telegram.manufacturing_type()
incomes_sheet = Parse_google_sheet(date_for_inspection)
incomes_sheet.connect_sheet()
incomes_sheet.seek_sheet()
worksheet_result = incomes_sheet.parse_sheet()
worksheet_result,finish_incomes = duplicate(worksheet_result, finish_incomes)
finall_check_inputs(worksheet_result, finish_incomes)
print('Час загрузки - ' + str((datetime.now() - start_time)))


#gs = gspread.service_account(filename='token.json')  # подключаем файл с ключами и пр.
#sh = gs.open_by_key('1oxu_S1yq7LPzJ9FurKM90e0FEDGKsKLUI28iTZ59anc')  # подключаем таблицу по ID
#sh = gs.open_by_key('1yusUPReWVPk4fr2jMtmbcxvMBoLyoJklp-B_YwAN-AM')  # подключаем таблицу по ID
#worksheet2 = sh.get_worksheet(1)
#worksheet = sh.get_worksheet(2)  # получаем второй лист
# получаем первый лист

#temporary_list = []
#with open("chats.csv", "r", errors="ignore") as f:
    #for line in f.readlines():
        #temporary_list.append(line.strip('\n'))
                       
#temporary_list_2 = []

#for i in range(len(temporary_list)):
    #if choose_date in temporary_list[i]:        
        #k = i
        #while True:
            #k+= 1
            #if '+00:00' in temporary_list[k]:
                #break
            #else:
                #b = temporary_list[k].strip('"')
                #b = b.split()
                #b = ''.join(b)
                #temporary_list_2.append(b)




#temporary_list_3 = []
#for i in temporary_list_2:
#    temporary_list_3.append([x.strip().lower() for x in i.rpartition('-')])
#
#print (temporary_list_3)
    



#temporary_incomes = {}
#for i in range(len(temporary_list_3)):
#    if temporary_list_3[i][1] != '' and temporary_list_3[i][0] not in temporary_incomes:
 #       temporary_incomes[temporary_list_3[i][0]] = temporary_list_3[i][2]
  #  elif temporary_list_3[i][0] in temporary_incomes:
   #     temporary_incomes[temporary_list_3[i][0]] = int(temporary_incomes[temporary_list_3[i][0]]) + int(temporary_list_3[i][2])

#print (temporary_incomes)

#for incomes_key, incomes_value in temporary_incomes.items():
#    if incomes_key not in conversion.values():
#        print(f'Неправильна ідентифікація! - {incomes_key}')
    
        


#incomes = {}
#for incomes_key, incomes_value in temporary_incomes.items():
#    for key, value in conversion.items():
#        if incomes_key == value:
#            incomes[key.lower()] = incomes_value
            

#print(incomes)
#print("Парсинг девайсів виконано.")

    
        
#first_cell = worksheet.find(date_for_inspection)
#first_column = get_column_letter(first_cell.col)
#last_cell = int(first_cell.col) + 3
#last_column = get_column_letter(last_cell)
#worksheet_result = [[0 for j in range(2)] for i in range(20)]
#worksheet_result = {}
#for i in range(281):
#for i in range(300):
    #res1 = worksheet.get(f"A{i+2}")
    #time.sleep(0.5)
#    res2 = worksheet.get(f"{first_column}{i+2}:{last_column}{i+2}")
#    sum_sheets = 0
#    print(res1)
#    if len(res2) > 0 and len(res1) > 0:
#        for k in res2[0]:
#            if k.isdigit():
#                sum_sheets += int(k)        
    #worksheet_result[i][1] = sum_sheets
    #worksheet_result[i][0].append(res1[0])
    #worksheet_result[i][0] = res1[0][0].lower()
    #worksheet_result[i][1] =sum_sheets
#        worksheet_result[res1[0][0].lower()] = sum_sheets

    





#k = 0
#for key, value in worksheet_result.items():
    #if key in finish_incomes:
        #if int(value) != int(finish_incomes[key]):
            #print(f'Невірно внесено - {key}. Повинно бути {finish_incomes[key]}, а внесено {value}')
            #k +=1 
    #elif key not in finish_incomes and int(value) != 0:
        #print(f"Внесено не туди. В рядку {key} повинно бути пусто")
        #k +=1
    #elif k == 0 :
        #print ("Обліковець не лажав, можна заварити йому чайок...Але наступного разу обов'язково напартачить")



